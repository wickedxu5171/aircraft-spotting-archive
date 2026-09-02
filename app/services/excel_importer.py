import re
from dataclasses import asdict, dataclass, replace
from datetime import date, datetime
from pathlib import Path
from typing import Iterator

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

from ..reference_data import AIRLINE_ALIASES


LEGACY_EVENT_GROUPS = (
    (1, 7, 8, 9),
    (2, 11, 12, 13),
    (3, 15, 16, 17),
    (4, 19, 20, 21),
)
COMPACT_EVENT_GROUPS = (
    (1, 6, 7, 8),
    (2, 10, 11, 12),
)
PLACEHOLDERS = {"", "/", "?", "？", "-", "—", "——"}
ROUTE_SEPARATOR = re.compile(r"\s*(?:—{2,}|–{2,}|-{2,}|→)\s*")
AIRCRAFT_TYPE_ALIASES = {
    "a380": "Airbus A380-841",
    "airbus a380": "Airbus A380-841",
    "a380-841": "Airbus A380-841",
    "airbus a380-841": "Airbus A380-841",
}


@dataclass(frozen=True)
class NormalizedSpottingEvent:
    source_row: int
    source_event_group: int
    airline_raw: str | None
    registration: str
    aircraft_type_raw: str | None
    aircraft_type: str | None
    aircraft_notes: str | None
    spotting_location_raw: str | None
    spotting_date: date | None
    flight_number: str | None
    route_text_original: str | None
    route_departure_raw: str | None
    route_arrival_raw: str | None
    quality_status: str
    quality_reasons: tuple[str, ...]

    def to_dict(self):
        result = asdict(self)
        result["spotting_date"] = (
            self.spotting_date.isoformat() if self.spotting_date else None
        )
        result["quality_reasons"] = list(self.quality_reasons)
        return result


def compact_text(value) -> str | None:
    if value is None:
        return None
    text = re.sub(r"\s+", " ", str(value)).strip()
    return text or None


def optional_text(value) -> str | None:
    text = compact_text(value)
    if text is None or text in PLACEHOLDERS:
        return None
    return text


def canonical_airline_name(value) -> str | None:
    name = optional_text(value)
    return AIRLINE_ALIASES.get(name, name)


def canonical_aircraft_type(value) -> str | None:
    """Preserve workbook detail while expanding known generic display labels."""
    aircraft_type = optional_text(value)
    if aircraft_type is None:
        return None
    return AIRCRAFT_TYPE_ALIASES.get(aircraft_type.casefold(), aircraft_type)


def normalize_registration(value) -> str | None:
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    text = compact_text(value)
    if text is None:
        return None
    return text.replace(" ", "").upper()


def normalize_date(value, workbook_epoch) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        return from_excel(value, workbook_epoch).date()
    if isinstance(value, str):
        value = value.strip()
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%d/%m/%Y"):
            try:
                return datetime.strptime(value, fmt).date()
            except ValueError:
                continue
    return None


def split_route(route: str | None) -> tuple[str | None, str | None]:
    if route is None:
        return None, None
    parts = ROUTE_SEPARATOR.split(route, maxsplit=1)
    if len(parts) != 2:
        return None, None
    return optional_text(parts[0]), optional_text(parts[1])


def _event_quality(
    spotting_date: date | None,
    spotting_location_raw: str | None,
    current_date: date,
) -> tuple[str, tuple[str, ...]]:
    reasons = []
    if spotting_date is None:
        reasons.append("missing_or_invalid_date")
    elif spotting_date > current_date:
        reasons.append("future_date")
    if spotting_location_raw and "/" in spotting_location_raw:
        reasons.append("ambiguous_spotting_location")
    return ("review" if reasons else "ready", tuple(reasons))


def repair_future_dates(
    events: list[NormalizedSpottingEvent],
    target_date: date,
    *,
    current_date: date | None = None,
) -> tuple[list[NormalizedSpottingEvent], int]:
    """Repair future Excel autofill years that share the target month/day.

    This is opt-in so a genuine future planning workbook is never changed
    silently. The original workbook remains untouched.
    """
    current_date = current_date or date.today()
    repaired = []
    repair_count = 0
    for event in events:
        should_repair = (
            event.spotting_date is not None
            and event.spotting_date > current_date
            and (event.spotting_date.month, event.spotting_date.day)
            == (target_date.month, target_date.day)
        )
        if not should_repair:
            repaired.append(event)
            continue
        quality_status, quality_reasons = _event_quality(
            target_date, event.spotting_location_raw, current_date
        )
        repaired.append(
            replace(
                event,
                spotting_date=target_date,
                quality_status=quality_status,
                quality_reasons=quality_reasons,
            )
        )
        repair_count += 1
    return repaired, repair_count


def _uses_compact_layout(sheet, workbook_epoch) -> bool:
    """Detect the corrected 12-column workbook without assuming a header row."""
    for row_number in range(1, min(sheet.max_row, 7) + 1):
        registration = normalize_registration(sheet.cell(row_number, 2).value)
        if not registration:
            continue
        if any(
            normalize_date(sheet.cell(row_number, column).value, workbook_epoch)
            for column in (6, 10)
        ):
            return True
    return False


def iter_spotting_events(
    workbook_path: str | Path,
    *,
    sheet_name: str = "航空",
    current_date: date | None = None,
) -> Iterator[NormalizedSpottingEvent]:
    current_date = current_date or date.today()
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        sheet = workbook[sheet_name]
        compact_layout = _uses_compact_layout(sheet, workbook.epoch)
        if compact_layout:
            min_row = 1
            max_col = 12
            event_groups = COMPACT_EVENT_GROUPS
            notes_index = None
            location_index = 3
        else:
            min_row = 8
            max_col = 21
            event_groups = LEGACY_EVENT_GROUPS
            notes_index = 3
            location_index = 4
        previous_location = None
        for row_number, row in enumerate(
            sheet.iter_rows(min_row=min_row, max_col=max_col, values_only=True),
            start=min_row,
        ):
            registration = normalize_registration(row[1])
            if not registration:
                previous_location = None
                continue

            airline_raw = canonical_airline_name(row[0])
            aircraft_type_raw = optional_text(row[2])
            aircraft_type = canonical_aircraft_type(aircraft_type_raw)
            aircraft_notes = (
                optional_text(row[notes_index]) if notes_index is not None else None
            )
            row_location = optional_text(row[location_index])
            if row_location:
                previous_location = row_location
            spotting_location_raw = row_location or previous_location

            for group_number, date_column, flight_column, route_column in event_groups:
                raw_date = row[date_column - 1]
                raw_flight = row[flight_column - 1]
                raw_route = row[route_column - 1]
                if raw_date is None and raw_flight is None and raw_route is None:
                    continue

                spotting_date = normalize_date(raw_date, workbook.epoch)
                flight_number = optional_text(raw_flight)
                if flight_number:
                    flight_number = flight_number.replace(" ", "").upper()
                route = optional_text(raw_route)
                departure, arrival = split_route(route)
                quality_status, quality_reasons = _event_quality(
                    spotting_date, spotting_location_raw, current_date
                )

                yield NormalizedSpottingEvent(
                    source_row=row_number,
                    source_event_group=group_number,
                    airline_raw=airline_raw,
                    registration=registration,
                    aircraft_type_raw=aircraft_type_raw,
                    aircraft_type=aircraft_type,
                    aircraft_notes=aircraft_notes,
                    spotting_location_raw=spotting_location_raw,
                    spotting_date=spotting_date,
                    flight_number=flight_number,
                    route_text_original=route,
                    route_departure_raw=departure,
                    route_arrival_raw=arrival,
                    quality_status=quality_status,
                    quality_reasons=quality_reasons,
                )
    finally:
        workbook.close()
